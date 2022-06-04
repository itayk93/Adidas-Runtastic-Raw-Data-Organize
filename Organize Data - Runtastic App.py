# Code by Itay Karkason
import pandas as pd
import glob
from datetime import date
from openpyxl import load_workbook

today = pd.Timestamp(date.today())
today_name = (str(today)[:-9]).replace("-", "")

# Get the path of the main folder for use
wb = load_workbook("main_folder_location.xlsx", data_only=True)
sh = wb["Sheet1"]
folder_location = sh["A2"].value
folder_location_output = folder_location + "Output/"

# Import all the json files from the input folder
all_files = glob.glob(folder_location + "Input/Exported Input Data/Sport-sessions/*.json")

li = []

for filename in all_files:
    df = pd.read_json(filename, lines=True)
    li.append(df)
    df['original_filename'] = filename.rsplit("/", 1)[-1]

frame = pd.concat(li, axis=0, ignore_index=True)

# convert dataframe to csv file
frame.to_csv(folder_location + "Input/Data Input Combined.csv", index=False)

data = pd.read_csv(folder_location + "Input/Data Input Combined.csv", parse_dates=['created_at'], dayfirst=False)

# Change the name of date column to 'Date'
data = data.rename(columns={"created_at": "date"})
data['date'] = pd.to_datetime(data['date'])
data['start_time'] = pd.to_datetime(data['start_time'])
data['end_time'] = pd.to_datetime(data['end_time'])

data_to_output = data[['id', 'date', 'start_time', 'end_time', 'distance', 'duration',
                       'elevation_gain', 'elevation_loss', 'average_speed', 'calories',
                       'duration_per_km', 'temperature', 'subjective_feeling_id',
                       'total_steps', 'original_filename']]

# Get the minimum average_speed to filter by
wb = load_workbook("min_average_speed.xlsx", data_only=True)
sh = wb["Sheet1"]
average_speed_cancel = sh["A2"].value
data_to_output = data_to_output[(data_to_output['average_speed'] > average_speed_cancel)]

# Filter the dates_to _cancel on the data
wb = load_workbook("dates_to _cancel.xlsx", data_only=True)
sh = wb["Sheet1"]
average_speed_cancel = sh["A2"].value

# Deleting data from dates that i need to cancel
dates_to_show = pd.read_excel("dates_to _cancel.xlsx")
start_date_to_cancel = pd.to_datetime(dates_to_show['start_date_to_cancel'])
end_date_to_cancel = pd.to_datetime(dates_to_show['end_date_to_cancel'])

dates_to_output_to_drop = pd.DataFrame(columns=data_to_output.columns)

for index, row in dates_to_show.iterrows():
    start_date_to_cancel = row['start_date_to_cancel']
    end_date_to_cancel = row['end_date_to_cancel']

    data_to_output_filtered_to_drop = data_to_output[
        (data_to_output['date'] >= start_date_to_cancel) &
        (data_to_output['date'] < end_date_to_cancel)
        ]
    print("The Data between " + str(start_date_to_cancel) + " to " + str(end_date_to_cancel) + " have been deleted")
    dates_to_output_to_drop = pd.concat([dates_to_output_to_drop, data_to_output_filtered_to_drop])

dates_to_output_to_drop_list = dates_to_output_to_drop['date'].values.tolist()
for date_to_cancel in dates_to_output_to_drop_list:
    data_to_output = data_to_output[(data_to_output['date'] != date_to_cancel)]

# Filter out data i decided to cancel
runs_to_cancel = pd.read_excel("data_to_cancel.xlsx")
id_to_cancel_list = runs_to_cancel['id'].to_list()
for id_to_cancel in id_to_cancel_list:
    date_of_run_canceled = data_to_output['date'][(data_to_output['id'] == id_to_cancel)]
    data_to_output = data_to_output[(data_to_output['id'] != id_to_cancel)]
    print("Run id - " + id_to_cancel + " from " + str(date_of_run_canceled)[6:-34] + " have been deleted")

# Taking just the starting time
data_to_output['start_time_in_hhmmss'] = data_to_output['start_time'].apply(
    lambda x: x.strftime('%H:%M:%S'))

# Sort by Date
data_to_output = data_to_output.sort_values(by="date")

list_for_int = ['elevation_gain', 'elevation_loss', 'calories', 'duration_per_km',
                'temperature', 'subjective_feeling_id', 'total_steps']
for i in list_for_int:
    data_to_output[i] = pd.to_numeric(data_to_output[i], downcast='integer')

# Export
data_to_output.to_excel(folder_location_output + "Data Runtastic Output.xlsx", index=False)
